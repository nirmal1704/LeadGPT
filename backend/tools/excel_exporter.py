"""
Excel export for LeadGPT.

Output sheets:
  Leads    — main data, sorted by score descending (colour-coded by category)
  Brief    — job brief fields and opportunity category definitions
  Metadata — export timestamp, lead counts, average score
"""
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_KNOWN_CATEGORY_COLORS: dict[str, str] = {
    "no_website":         "C6EFCE",
    "broken_link":        "FFC7CE",
    "insecure_http":      "FFEB9C",
    "no_mobile":          "FCE4D6",
    "outdated_design":    "FFFF99",
    "no_social_presence": "E2EFDA",
    "slow_load":          "FFF2CC",
    "no_online_booking":  "DAE8FC",
    "no_google_listing":  "E1D5E7",
    "no_ecommerce":       "D5E8D4",
    "no_whatsapp_link":   "DAE8FC",
    "functional":         "FFFFFF",
    "unknown":            "F2F2F2",
}

_DYNAMIC_PALETTE = [
    "D9E1F2", "FCE4D6", "E2EFDA", "FFF2CC", "F4CCCC",
    "CFE2F3", "EAD1DC", "D9D2E9", "D0E0E3", "FFF3E0",
]

_COLUMNS = [
    ("rank",                 "Rank"),
    ("name",                 "Business Name"),
    ("phone",                "Phone"),
    ("address",              "Address"),
    ("website_url",          "Website"),
    ("opportunity_category", "Category"),
    ("opportunity_score",    "Score"),
    ("pitch_angle",          "Pitch Angle"),
    ("is_low_priority",      "Low Priority?"),
    ("social_media_url",     "Social Media"),
    ("is_https",             "HTTPS?"),
    ("has_mobile_viewport",  "Mobile Friendly?"),
    ("h1_tags",              "H1 Tags"),
    ("meta_description",     "Meta Description"),
    ("internal_links_count", "Internal Links"),
    ("source_url",           "Source URL"),
]

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT   = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def export_to_excel(
    leads: list[dict],
    file_path: str,
    job_brief: dict | None = None,
    opportunity_categories: list[str] | None = None,
    category_rules: dict | None = None,
) -> str:
    """Write leads to an Excel file and return the file path."""
    job_brief = job_brief or {}
    opportunity_categories = opportunity_categories or []
    category_rules = category_rules or {}

    color_map = _build_color_map(opportunity_categories)
    sorted_leads = sorted(
        leads,
        key=lambda x: (not x.get("is_low_priority", False), -int(x.get("opportunity_score", 0)))
    )

    wb = Workbook()
    _write_leads_sheet(wb, sorted_leads, color_map)
    _write_brief_sheet(wb, job_brief, opportunity_categories, category_rules)
    _write_metadata_sheet(wb, leads, job_brief)

    wb.save(file_path)
    return file_path


def _build_color_map(opportunity_categories: list[str]) -> dict[str, str]:
    color_map = dict(_KNOWN_CATEGORY_COLORS)
    palette_idx = 0
    for code in opportunity_categories:
        if code not in color_map:
            color_map[code] = _DYNAMIC_PALETTE[palette_idx % len(_DYNAMIC_PALETTE)]
            palette_idx += 1
    return color_map


def _write_leads_sheet(wb: Workbook, leads: list[dict], color_map: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "Leads"

    for col_idx, (_, header) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = _THIN_BORDER

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, lead in enumerate(leads, start=2):
        category = str(lead.get("opportunity_category", "unknown"))
        fill = PatternFill(
            start_color=color_map.get(category, "F2F2F2"),
            end_color=color_map.get(category, "F2F2F2"),
            fill_type="solid",
        )

        for col_idx, (field, _) in enumerate(_COLUMNS, start=1):
            if field == "rank":
                value: Any = row_idx - 1
            elif field == "h1_tags":
                raw = lead.get("h1_tags")
                value = ", ".join(str(t) for t in raw[:3]) if isinstance(raw, list) else (str(raw) if raw else "")
            elif field == "is_https":
                site = lead.get("site_data", {})
                value = "Yes" if site.get("is_https") or lead.get("website_url", "").startswith("https://") else "No"
            elif field == "has_mobile_viewport":
                value = "Yes" if lead.get("site_data", {}).get("has_mobile_viewport") else "No"
            elif field == "meta_description":
                site = lead.get("site_data", {})
                value = site.get("meta_description", lead.get("meta_description", ""))
            elif field == "internal_links_count":
                site = lead.get("site_data", {})
                value = site.get("internal_links_count", lead.get("internal_links_count", ""))
            elif field == "is_low_priority":
                value = "Yes" if lead.get("is_low_priority") else "No"
            else:
                value = lead.get(field, "") or ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(field in ("pitch_angle", "address")))

    _auto_width(ws, min_width=10, max_width=60)


def _write_brief_sheet(
    wb: Workbook,
    brief: dict,
    opportunity_categories: list[str],
    category_rules: dict,
) -> None:
    ws = wb.create_sheet(title="Brief")

    rows = [
        ("Who we are",      brief.get("user_context", "")),
        ("Offering",        brief.get("offering", "")),
        ("Lead category",   brief.get("lead_category", "")),
        ("Location",        brief.get("location", "")),
        ("Leads requested", brief.get("lead_count", "")),
        ("Notes",           brief.get("additional_notes", "")),
        ("", ""),
        ("Opportunity categories", "Definition"),
    ]
    for code in opportunity_categories:
        rows.append((code, category_rules.get(code, "")))

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font   = Font(name="Calibri", size=10)

    for row_idx, (key, value) in enumerate(rows, start=1):
        for col_idx, v in enumerate([key, value], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = header_font if row_idx == 8 else body_font
            if row_idx == 8:
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _auto_width(ws, min_width=20, max_width=70)


def _write_metadata_sheet(wb: Workbook, leads: list[dict], brief: dict) -> None:
    ws = wb.create_sheet(title="Metadata")

    rows = [
        ("export_timestamp", datetime.now(timezone.utc).isoformat()),
        ("total_leads",      len(leads)),
        ("leads_requested",  brief.get("lead_count", "")),
        ("lead_category",    brief.get("lead_category", "")),
        ("location",         brief.get("location", "")),
        ("high_priority",    sum(1 for l in leads if not l.get("is_low_priority"))),
        ("low_priority",     sum(1 for l in leads if l.get("is_low_priority"))),
        ("avg_score",        round(sum(int(l.get("opportunity_score", 0)) for l in leads) / max(len(leads), 1), 2)),
    ]

    for row_idx, (key, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=key).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row_idx, column=2, value=value).font = Font(name="Calibri", size=10)

    _auto_width(ws, min_width=15, max_width=50)


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)
